
from openpyxl import load_workbook
import os

file_path = r'../data/raw/ficha_caracterizacao.xlsx'

print(f"Tentando ler: {file_path}")

try:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"Abas disponíveis: {wb.sheetnames}")
    
    # Procura 'informações'
    target_sheet = None
    for sheet in wb.sheetnames:
        if 'inform' in sheet.lower():
            target_sheet = sheet
            print(f">>> Aba alvo encontrada: {target_sheet}")
            break
            
    if target_sheet:
        ws = wb[target_sheet]
        print(f"\n--- Conteúdo da aba '{target_sheet}' (primeiras 50 linhas) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 50: break
            # Imprime apenas células com texto não vazio
            cleaned_row = [cell for cell in row if cell is not None]
            if cleaned_row:
                print(f"Linha {i+1}: {cleaned_row}")
    else:
        print("Aba de 'INFORMAÇÕES' não encontrada explicitamente. Mostrando a primeira aba.")
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 20: break
            cleaned_row = [cell for cell in row if cell is not None]
            if cleaned_row:
                print(f"Linha {i+1}: {cleaned_row}")

except Exception as e:
    print(f"Erro ao ler Excel com openpyxl: {e}")
